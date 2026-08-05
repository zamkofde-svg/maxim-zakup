"""Универсальный парсер прайс-листов поставщиков (xlsx + docx-таблицы).

Логика per-row (не per-column), чтобы пережить разные форматы:
- строка-товар = есть НАЗВАНИЕ + положительная ЦЕНА (число)
- строка-категория = есть текст, но нет цены → пропускаем
- НАЗВАНИЕ = самый левый непустой текст, который НЕ страна/единица/валюта/число
- колонку «Страна/происхождение» вычисляем из шапки, чтобы не спутать с названием
"""
from __future__ import annotations
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_UNIT_RE = re.compile(r"^\s*(кг|шт|л|уп|упак|гр|г|мл|ед|бут|пач|бан|короб|ящик)\.?\s*$", re.I)
_CUR_RE = re.compile(r"^\s*(руб|₽|р)\.?\s*$", re.I)


_MAX_PRICE = 500000   # выше — это не цена, а номер счёта/телефона/ИНН


def _to_price(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if 0 < v <= _MAX_PRICE else None
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    # если в исходнике много цифр подряд (счёт/тел) — не цена
    if re.sub(r"\D", "", s) and len(re.sub(r"\D", "", s)) > 7:
        return None
    s = re.sub(r"[^0-9.]", "", s)
    if not s:
        return None
    try:
        f = float(s)
        return f if 0 < f <= _MAX_PRICE else None
    except ValueError:
        return None


def _is_unit(v):
    return isinstance(v, str) and bool(_UNIT_RE.match(v))


def _is_currency(v):
    return isinstance(v, str) and bool(_CUR_RE.match(v))


def _norm_unit(v):
    """КГ/кг/л → kg_or_l; шт/уп/... → pkg."""
    if not v:
        return None
    s = str(v).strip().lower()
    if s.startswith(("кг", "л", "гр", "г", "мл")):
        return "kg_or_l"
    return "pkg"


def _read_rows_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    rows = []
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    return rows


def _read_rows_xls(path):
    """Старый формат .xls (BIFF/OLE2) — читаем через xlrd."""
    import xlrd
    wb = xlrd.open_workbook(str(path))
    rows = []
    for ws in wb.sheets():
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                v = ws.cell_value(r, c)
                row.append(v if v not in ("", None) else None)
            rows.append(row)
    return rows


def _read_rows_docx(path):
    z = zipfile.ZipFile(path)
    root = ET.fromstring(z.read("word/document.xml").decode("utf-8"))
    rows = []
    for tbl in root.iter(_W + "tbl"):
        for tr in tbl.iter(_W + "tr"):
            cells = []
            for tc in tr.iter(_W + "tc"):
                txt = "".join(t.text or "" for t in tc.iter(_W + "t")).strip()
                cells.append(txt if txt else None)
            rows.append(cells)
    return rows


def _find_country_col(rows):
    """Индекс колонки со страной/происхождением из шапки (или None)."""
    for row in rows[:20]:
        for i, c in enumerate(row):
            if isinstance(c, str) and re.search(r"стран|происхожд", c, re.I):
                return i
    return None


def parse_pricelist(path):
    path = Path(path)
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        rows = _read_rows_xlsx(path)
    elif ext == ".xls":
        rows = _read_rows_xls(path)
    elif ext == ".docx":
        rows = _read_rows_docx(path)
    else:
        raise ValueError(f"Неподдерживаемый формат: {ext}")

    country_col = _find_country_col(rows)
    items = []
    for row in rows:
        if not row:
            continue
        # цена = первое число >0 среди ячеек
        price = None
        price_idx = None
        for i, c in enumerate(row):
            p = _to_price(c)
            if p is not None:
                price, price_idx = p, i
                break
        if price is None:
            continue  # категория/шапка/пусто
        # единица
        unit = None
        for c in row:
            if _is_unit(c):
                unit = _norm_unit(c)
                break
        # название = самый левый текст, не страна/единица/валюта/число/до цены
        name = None
        for i, c in enumerate(row):
            if not isinstance(c, str) or not c.strip():
                continue
            if i == country_col:
                continue
            if _is_unit(c) or _is_currency(c) or _to_price(c) is not None:
                continue
            if len(c.strip()) < 2:
                continue
            name = re.sub(r"\s+", " ", c).strip()
            break
        if not name:
            continue
        # отсечём явный мусор в названии (только цифры/№)
        if re.fullmatch(r"[\d.,№\s]+", name):
            continue
        items.append({"name": name, "price": price, "unit": unit})
    return items


# ============ СОПОСТАВЛЕНИЕ С МАСТЕР-МАТРИЦЕЙ ============

_COUNTRIES = {
    "египет", "марокко", "россия", "рф", "азербайджан", "узбекистан", "казахстан",
    "израиль", "турция", "иран", "китай", "сербия", "эквадор", "юар", "грузия",
    "армения", "молдова", "испания", "италия", "греция", "тюмень", "азер", "узб",
}
_STOP = {
    "свежий", "свежие", "свежая", "свежее", "урожай", "в", "вакуумной", "упаковке",
    "упаковка", "крупный", "крупное", "крупная", "крупные", "мелкий", "мелкое",
    "отборный", "собственного", "производства", "св", "ур", "кг", "шт", "го",
    "сорт", "кор", "ящик", "стакан", "пучок", "вес", "весовой", "весовая",
}
_ENDINGS = ("ами", "ями", "ов", "ев", "ые", "ий", "ая", "ое", "ы", "и", "а", "я", "е", "ь", "ой", "ую", "ым")


def _stem(t: str) -> str:
    for suf in _ENDINGS:
        if len(t) > 4 and t.endswith(suf):
            return t[:-len(suf)]
    return t


def match_tokens(s: str) -> set:
    """Токены для fuzzy-сопоставления: без чисел/единиц/стоп-слов/стран, со стеммингом."""
    s = str(s or "").lower().replace("ё", "е")
    s = re.sub(r"\d+[.,]?\d*\s*(кг|г|мл|л|шт|уп|гр)\b", " ", s)
    s = re.sub(r"[^a-zа-я0-9 ]+", " ", s)
    return {_stem(t) for t in s.split() if t and t not in _STOP and t not in _COUNTRIES and len(t) > 1}


def alias_key(raw: str) -> str:
    """Ключ для запоминания сопоставления: то же название из прайса того же
    поставщика должно давать тот же ключ (lowercase, схлопнутые пробелы, ё→е)."""
    s = str(raw or "").lower().replace("ё", "е")
    return re.sub(r"\s+", " ", s).strip()


def build_master_index(master: list) -> list:
    """master: [{id, name, category, category_id, unit_type}] → [(item, tokenset)]."""
    return [(m, match_tokens(m["name"])) for m in master]


def match_name(name: str, index: list, topn: int = 4) -> list:
    """Возвращает топ-N кандидатов [(master_item, score)] по убыванию score (0..1)."""
    q = match_tokens(name)
    if not q:
        return []
    scored = []
    for m, mt in index:
        if not mt:
            continue
        inter = len(q & mt)
        if not inter:
            continue
        union = len(q | mt)
        score = inter / union if union else 0
        if q <= mt or mt <= q:
            # ПОЛНОЕ вхождение одного набора в другой. «Точно» (0.85) даём ТОЛЬКО если
            # покрыта бо́льшая часть большего набора — т.е. совпадение почти полное.
            # Иначе «сахар»→«Сахар ванильный», «перец»→«Перец красный», «нут»→«Крупа Нут…»
            # ложно попадали в «точно» и заливались не туда. Частичное вхождение → 0.5
            # («проверьте»), закупщик решает сам.
            coverage = inter / max(len(q), len(mt))
            score = max(score, 0.85 if coverage >= 0.6 else 0.5)
        scored.append((m, score))
    scored.sort(key=lambda x: -x[1])
    return scored[:topn]


if __name__ == "__main__":
    import sys
    files = sys.argv[1:] or []
    for f in files:
        items = parse_pricelist(f)
        print(f"\n===== {Path(f).name}: извлечено {len(items)} позиций =====")
        for it in items[:12]:
            print(f"  {it['name'][:45]:<47} | {it['price']:>8} | {it['unit']}")
        if len(items) > 12:
            print(f"  … ещё {len(items)-12}")
