"""
ETL-импортёр: Drive snapshot → SQLite.

Идемпотентный: запускать сколько угодно раз. Цены меняются — пишем в history,
обновляем quotes. Мастер-номенклатура — синхронизируется (insert/update/delete).
"""
from __future__ import annotations
import sys
import re
from datetime import datetime, date
from pathlib import Path
from typing import Iterable

# Чтобы импортировать наши etl-парсеры
sys.path.insert(0, str(Path(__file__).parent.parent / "etl"))

from openpyxl import load_workbook
from sqlalchemy import select, delete, func
from sqlalchemy.orm import Session

from backend.db import SessionLocal, init_db
from backend.models import (
    Category, AccountingSystem, Restaurant,
    Supplier, SupplierAlias,
    ProductMaster, AccountingAlias,
    PriceQuote, PriceHistory, PriceChange,
    PurchaseFact, Deviation,
    ImportRun, UnmappedItem,
)

from parse_supplier_matrix import parse as parse_matrix
from parse_mapping import parse as parse_mapping
from parse_fact_iiko import parse_iiko
from parse_fact_storehouse import parse_storehouse


import os
DRIVE_DIR = Path(os.environ.get("DRIVE_SYNC_DIR", str(Path(__file__).parent.parent / "sample-data" / "drive-sync")))
SAMPLES_DIR = Path(__file__).parent.parent / "sample-data"

MASTER_FILE = "Матрица(для изменения позиций).xlsx"
MAPPING_FILE = "Карта сопоставлений.xlsx"
FACTS_DIR = DRIVE_DIR / "facts"  # подпапка для выгрузок iiko/SH

# Служебные файлы — НЕ матрицы поставщиков (от старого разработчика, не парсим)
NON_SUPPLIER_FILES = {
    MASTER_FILE, MAPPING_FILE,
    "Топ 2.xlsx", "Сопоставление.xlsx", "Сводная.xlsx",
}

# Префиксы названий организаций — то, что начинается так, считаем матрицей поставщика
SUPPLIER_PREFIXES = ("ООО ", "АО ", "ИП ", "ПАО ", "ЗАО ", "АО ")


def discover_supplier_matrices(drive_dir: Path) -> list[tuple[str, Path]]:
    """Авто-сканирование папки drive-sync: все .xlsx с названием поставщика."""
    result = []
    for path in sorted(drive_dir.glob("*.xlsx")):
        name = path.name
        if name in NON_SUPPLIER_FILES:
            continue
        if not any(name.startswith(p) for p in SUPPLIER_PREFIXES):
            continue
        # Имя поставщика = имя файла без .xlsx
        supplier_name = name[:-len(".xlsx")]
        result.append((supplier_name, path))
    return result

def detect_fact_format(path: Path) -> str:
    """Авто-детект формата выгрузки факта по содержимому файла.
    StoreHouse экспортит XML с расширением .xls, iiko — настоящий xlsx.
    """
    with open(path, "rb") as f:
        head = f.read(200)
    if head.startswith(b"<?xml") or b"spreadsheet" in head[:200].lower():
        return "storehouse"
    return "iiko"


def discover_fact_files() -> list[dict]:
    """Сканирует подпапку drive-sync/facts/ и возвращает список файлов с auto-detected форматом."""
    if not FACTS_DIR.exists():
        return []
    result = []
    for p in sorted(FACTS_DIR.glob("*")):
        if not (p.suffix.lower() in (".xls", ".xlsx")):
            continue
        fmt = detect_fact_format(p)
        # система учёта — пока всем SH; в будущем — по имени файла можно различать
        result.append({"path": p, "format": fmt, "system": "SH"})
    return result

# Категории и юнит-тайп для каждой
CATEGORY_UNITS = {
    "Сыры": "kg_or_l", "Молочка": "kg_or_l",
    "Мясо": "pkg", "Рыба и морепродукты": "pkg", "макароны": "pkg",
    "Шоколад": "pkg", "Яйцо": "pkg", "Ягода см": "pkg",
    "Бакалея": "pkg", "Консервация": "pkg", "Мукасмеси": "pkg",
    "Овощифрукты": "kg_or_l",
}

SYSTEMS = ["SH", "Chees", "TEHNIKUM", "Sorrento"]


# ============ УТИЛИТЫ ============

def normalize(s: str | None) -> str:
    if not s:
        return ""
    s = str(s).lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s.replace('"', "").replace("«", "").replace("»", "")


def get_or_create(db: Session, model, defaults: dict | None = None, **kwargs):
    """Стандартный get-or-create. Возвращает (instance, created)."""
    obj = db.execute(select(model).filter_by(**kwargs)).scalar_one_or_none()
    if obj:
        return obj, False
    params = {**kwargs, **(defaults or {})}
    obj = model(**params)
    db.add(obj)
    db.flush()
    return obj, True


# ============ ШАГИ ИМПОРТА ============

def seed_static(db: Session):
    """Категории + системы учёта."""
    for name, unit in CATEGORY_UNITS.items():
        get_or_create(db, Category, name=name, defaults={"unit_type": unit})
    for s in SYSTEMS:
        get_or_create(db, AccountingSystem, name=s)
    db.flush()


def import_master_matrix(db: Session, path: Path) -> int:
    """Загружает мастер-матрицу как products_master."""
    wb = load_workbook(path, data_only=True)
    count = 0
    for sheet_name in wb.sheetnames:
        cat = db.execute(select(Category).filter_by(name=sheet_name)).scalar_one_or_none()
        if not cat:
            # Категория не описана в наших CATEGORY_UNITS — создадим с default
            cat = Category(name=sheet_name, unit_type="pkg")
            db.add(cat)
            db.flush()

        ws = wb[sheet_name]
        # Унифицированная логика: A = name, B (если "Ед.изм") = unit_label
        is_veg_format = sheet_name == "Овощифрукты"
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name or not str(name).strip():
                continue
            name = re.sub(r"\s+", " ", str(name)).strip()
            if name == "Наименование товара" or name.startswith("Наименование"):
                continue

            unit_label = None
            if is_veg_format:
                v = ws.cell(r, 2).value
                if v:
                    unit_label = str(v).strip()

            obj, created = get_or_create(
                db, ProductMaster,
                category_id=cat.id, name_normalized=normalize(name),
                defaults={"name": name, "unit_label": unit_label},
            )
            if not created:
                # обновим имя на актуальное (вдруг подправили регистр/пробелы)
                if obj.name != name:
                    obj.name = name
            count += 1
    return count


def import_mapping(db: Session, path: Path):
    """
    Карта сопоставлений → accounting_aliases + supplier_aliases.
    Сначала чистим — это справочник, не история.
    """
    records, sup_aliases = parse_mapping(path)

    # Поставщики из карты
    canonical_suppliers = {sa.canonical for sa in sup_aliases}
    for cname in canonical_suppliers:
        sup, _ = get_or_create(
            db, Supplier,
            name_normalized=normalize(cname),
            defaults={"name": cname},
        )
    db.flush()

    # Аккаунтинг-алиасы: чистим и наполняем заново
    db.execute(delete(AccountingAlias))
    sys_cache = {s.name: s for s in db.execute(select(AccountingSystem)).scalars()}

    # Маппим (supplier_name из карты) → product_master_id
    # supplier_name в карте = имя как в матрицах поставщиков
    pm_by_normname: dict[str, list[ProductMaster]] = {}
    for pm in db.execute(select(ProductMaster)).scalars():
        pm_by_normname.setdefault(pm.name_normalized, []).append(pm)

    added = 0
    unresolved = 0
    for r in records:
        system = sys_cache.get(r.system)
        if not system:
            continue
        matches = pm_by_normname.get(normalize(r.supplier_name), [])
        if not matches:
            unresolved += 1
            continue
        # если позиция в разных категориях — берём первую (фронт может уточнять)
        pm = matches[0]
        db.add(AccountingAlias(
            product_master_id=pm.id,
            system_id=system.id,
            name=r.system_name,
            name_normalized=normalize(r.system_name),
        ))
        added += 1

    # Алиасы поставщиков: чистим и наполняем
    db.execute(delete(SupplierAlias))
    sup_cache = {s.name_normalized: s for s in db.execute(select(Supplier)).scalars()}

    seen_pairs: set[tuple[int, str]] = set()
    for sa in sup_aliases:
        canon = sup_cache.get(normalize(sa.canonical))
        if not canon:
            continue
        norm_alias = normalize(sa.alias)
        key = (canon.id, norm_alias)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        db.add(SupplierAlias(
            supplier_id=canon.id,
            alias=sa.alias,
            alias_normalized=norm_alias,
        ))
    db.flush()
    return {"aliases_added": added, "unresolved": unresolved}


def import_supplier_matrix(db: Session, path: Path, supplier_name: str) -> dict:
    """Парсит матрицу одного поставщика → price_quotes (с архивированием изменений)."""
    quotes = parse_matrix(path, supplier_name)

    sup, _ = get_or_create(
        db, Supplier,
        name_normalized=normalize(supplier_name),
        defaults={"name": supplier_name},
    )

    # Кеш мастер-позиций по (категория, normalized name)
    cat_cache = {c.name: c for c in db.execute(select(Category)).scalars()}
    pm_cache: dict[tuple[int, str], ProductMaster] = {}
    for pm in db.execute(select(ProductMaster)).scalars():
        pm_cache[(pm.category_id, pm.name_normalized)] = pm

    # Кеш уже обработанных (supplier_id, product_master_id) в этом проходе —
    # чтобы не вставить дубль если в матрице две одинаковые позиции
    seen_pairs: set[tuple[int, int]] = set()

    inserted, updated, history, unmatched, snapshot_saved = 0, 0, 0, 0, 0
    now = datetime.utcnow()

    # Snapshot policy: пишем срез цен раз в сутки даже если они не менялись,
    # чтобы потом матчить факт по «цене на дату закупки».
    # Проверяем — есть ли price_history запись за последние 20 часов для этого поставщика
    from datetime import timedelta
    snapshot_threshold = now - timedelta(hours=20)
    last_snap_ts = db.scalar(
        select(func.max(PriceHistory.captured_at))
        .where(PriceHistory.supplier_id == sup.id)
    )
    need_snapshot = (last_snap_ts is None) or (last_snap_ts < snapshot_threshold)

    for q in quotes:
        cat = cat_cache.get(q.category)
        if not cat:
            unmatched += 1
            continue
        pm = pm_cache.get((cat.id, normalize(q.product)))
        if not pm:
            unmatched += 1
            continue

        # Dedup: если эту пару (sup, pm) уже обрабатывали в этом проходе — пропускаем
        if (sup.id, pm.id) in seen_pairs:
            continue
        seen_pairs.add((sup.id, pm.id))

        existing = db.execute(
            select(PriceQuote).filter_by(supplier_id=sup.id, product_master_id=pm.id)
        ).scalar_one_or_none()

        if existing:
            if existing.unit_price != q.unit_price:
                # архивируем старое (изменение)
                db.add(PriceHistory(
                    supplier_id=sup.id, product_master_id=pm.id,
                    unit_price=existing.unit_price, captured_at=existing.captured_at,
                ))
                # Фиксируем РЕАЛЬНОЕ изменение цены в audit-таблице
                delta_pct = ((q.unit_price - existing.unit_price) / existing.unit_price * 100) if existing.unit_price else 0
                db.add(PriceChange(
                    supplier_id=sup.id, product_master_id=pm.id,
                    old_price=existing.unit_price, new_price=q.unit_price,
                    delta_pct=delta_pct, changed_at=now,
                ))
                existing.unit_price = q.unit_price
                existing.pkg_net = q.pkg_net
                existing.pkg_gross = q.pkg_gross
                existing.supplier_comment = q.supplier_comment
                existing.captured_at = now
                history += 1
            elif need_snapshot:
                # Цена та же, но снапшот за сегодня нужен — для исторической сверки
                db.add(PriceHistory(
                    supplier_id=sup.id, product_master_id=pm.id,
                    unit_price=existing.unit_price, captured_at=now,
                ))
                snapshot_saved += 1
            # Комментарий может меняться отдельно от цены — обновляем всегда, если отличается
            if existing.supplier_comment != q.supplier_comment:
                existing.supplier_comment = q.supplier_comment
            updated += 1
        else:
            db.add(PriceQuote(
                supplier_id=sup.id, product_master_id=pm.id,
                unit_price=q.unit_price, unit_type=q.unit_type,
                pkg_net=q.pkg_net, pkg_gross=q.pkg_gross,
                supplier_comment=q.supplier_comment, captured_at=now,
            ))
            # Сразу пишем эту же цену в history — это «известная цена с такой-то даты»
            db.add(PriceHistory(
                supplier_id=sup.id, product_master_id=pm.id,
                unit_price=q.unit_price, captured_at=now,
            ))
            snapshot_saved += 1
            inserted += 1

    # === УДАЛЕНИЕ ИСЧЕЗНУВШИХ ЦЕН ===
    # Если в БД есть PriceQuote по (sup, pm), а в текущей матрице его уже нет
    # (поставщик стёр цену из ячейки или мы удалили позицию при prune) —
    # удаляем эту цену из БД. PriceHistory не трогаем (это исторический архив).
    db.flush()
    seen_pm_ids = {pm_id for _, pm_id in seen_pairs}
    stale = db.execute(
        select(PriceQuote).where(PriceQuote.supplier_id == sup.id)
    ).scalars().all()
    removed = 0
    for old in stale:
        if old.product_master_id not in seen_pm_ids:
            db.delete(old)
            removed += 1

    return {
        "inserted": inserted, "updated": updated, "history_changes": history,
        "snapshots": snapshot_saved, "unmatched": unmatched, "removed_stale": removed,
    }


def import_facts(db: Session) -> dict:
    """Загружает выгрузки факта и сразу матчит → deviations.

    ВАЖНО: сохраняем указанные шефами причины ДО wipe и восстанавливаем
    по натуральному ключу (дата, товар, поставщик, ресторан).
    """
    # 1. СНАЧАЛА сохраняем все указанные причины (по натуральному ключу)
    saved_reasons = {}  # (date, raw_product, raw_supplier, raw_restaurant) → (reason_text, reason_category)
    for dev, pf in db.execute(
        select(Deviation, PurchaseFact)
        .join(PurchaseFact, Deviation.purchase_fact_id == PurchaseFact.id)
        .where((Deviation.reason_text != None) | (Deviation.reason_category != None))
    ).all():
        key = (pf.date.isoformat(), pf.raw_product, pf.raw_supplier, pf.raw_restaurant or "")
        saved_reasons[key] = (dev.reason_text, dev.reason_category)
    if saved_reasons:
        print(f"   → сохранено {len(saved_reasons)} причин для восстановления после пересчёта")

    # 2. Чистим старое (для MVP — переимпорт всегда полный)
    db.execute(delete(Deviation))
    db.execute(delete(PurchaseFact))
    db.flush()

    # Кеши для матчинга
    sup_alias_cache = {
        sa.alias_normalized: sa.supplier_id
        for sa in db.execute(select(SupplierAlias)).scalars()
    }
    sup_name_cache = {
        s.name_normalized: s.id
        for s in db.execute(select(Supplier)).scalars()
    }
    sys_cache = {s.name: s for s in db.execute(select(AccountingSystem)).scalars()}
    acct_alias_cache: dict[tuple[int, str], int] = {}
    for aa in db.execute(select(AccountingAlias)).scalars():
        acct_alias_cache.setdefault((aa.system_id, aa.name_normalized), aa.product_master_id)

    # === Цены ПО ДАТАМ — берём ВСЁ из price_history + текущие из price_quotes
    # Структура: history_by_pair[(supplier_id, pm_id)] = sorted list of (date, price)
    history_by_pair: dict[tuple[int, int], list[tuple[datetime, float]]] = {}
    for ph in db.execute(select(PriceHistory).order_by(PriceHistory.captured_at)).scalars():
        history_by_pair.setdefault((ph.supplier_id, ph.product_master_id), []).append(
            (ph.captured_at, ph.unit_price)
        )
    # Текущие цены тоже добавим — это «цена с такой-то даты до сих пор»
    for pq in db.execute(select(PriceQuote)).scalars():
        history_by_pair.setdefault((pq.supplier_id, pq.product_master_id), []).append(
            (pq.captured_at, pq.unit_price)
        )
    # Пере-сортируем каждый список по дате (на всякий случай)
    for key in history_by_pair:
        history_by_pair[key].sort(key=lambda x: x[0])

    def price_on_date(supplier_id: int, pm_id: int, target_date: date) -> Optional[float]:
        """Возвращает цену поставщика на товар на конкретную дату:
        - последний снапшот с captured_at <= target_date
        - если все снапшоты позже target_date — берём самый ранний (приближение, ничего лучше нет)
        - если данных вообще нет — None"""
        records = history_by_pair.get((supplier_id, pm_id))
        if not records:
            return None
        # ищем последнюю запись с датой <= target_date
        target_dt = datetime(target_date.year, target_date.month, target_date.day, 23, 59, 59)
        best = None
        for ts, price in records:
            if ts <= target_dt:
                best = price
            else:
                break
        if best is not None:
            return best
        # все снапшоты ПОЗЖЕ target_date — берём самый ранний (приближение)
        return records[0][1]

    def top2_on_date(pm_id: int, target_date: date) -> tuple:
        """Возвращает (top1_sup, top1_price, top2_sup, top2_price) на дату.
        Учитываются только те поставщики у которых была цена на эту дату."""
        candidates = []
        for (sup_id, pm), records in history_by_pair.items():
            if pm != pm_id:
                continue
            price = price_on_date(sup_id, pm_id, target_date)
            if price is not None:
                candidates.append((sup_id, price))
        if not candidates:
            return (None, None, None, None)
        candidates.sort(key=lambda x: x[1])
        t1 = candidates[0]
        t2 = candidates[1] if len(candidates) > 1 else (None, None)
        return (t1[0], t1[1], t2[0], t2[1])

    # Рестораны — собираем из SH-выгрузки на лету
    restaurants_cache: dict[str, Restaurant] = {
        r.sh_code: r for r in db.execute(select(Restaurant)).scalars() if r.sh_code
    }

    total_facts = 0
    bucket = {}

    fact_files = discover_fact_files()
    # Дополнительно — sample-data для совместимости (на проде они не будут существовать)
    legacy = [
        {"path": SAMPLES_DIR / "iiko_or_sh_1.xls",  "format": "storehouse", "system": "SH"},
        {"path": SAMPLES_DIR / "iiko_or_sh_2.xlsx", "format": "iiko",       "system": "SH"},
    ]
    for spec in legacy:
        if spec["path"].exists():
            fact_files.append(spec)
    print(f"→ найдено файлов факта: {len(fact_files)}")

    for spec in fact_files:
        if not spec["path"].exists():
            continue
        try:
            if spec["format"] == "iiko":
                facts = parse_iiko(spec["path"])
            else:
                facts = parse_storehouse(spec["path"])
            print(f"   → {spec['path'].name} ({spec['format']}): {len(facts)} записей")
        except Exception as e:
            print(f"   ⚠ {spec['path'].name} ({spec['format']}): ошибка {type(e).__name__}: {e}")
            continue

        system = sys_cache.get(spec["system"])

        for f in facts:
            # 1) Поставщик: алиас → канон, либо новый
            norm_sup = normalize(f.supplier)
            sup_id = sup_alias_cache.get(norm_sup) or sup_name_cache.get(norm_sup)
            is_internal = "Цех" in f.supplier or "Производство" in f.supplier
            if not sup_id:
                new_sup, _ = get_or_create(
                    db, Supplier, name_normalized=norm_sup,
                    defaults={"name": f.supplier, "is_internal": is_internal}
                )
                sup_id = new_sup.id
                sup_name_cache[norm_sup] = sup_id

            # 2) Ресторан (только SH)
            rest = None
            if f.restaurant:
                code = f.restaurant.split("/")[0].strip()
                rest = restaurants_cache.get(code)
                if not rest:
                    rest = Restaurant(name=code, sh_code=code)
                    db.add(rest)
                    db.flush()
                    restaurants_cache[code] = rest

            # 3) Мастер-позиция через карту сопоставлений
            pm_id = acct_alias_cache.get((system.id, normalize(f.product))) if system else None

            # 4) Парсим дату
            try:
                d = date.fromisoformat(f.date)
            except (ValueError, TypeError):
                continue

            pf = PurchaseFact(
                source=f.source, source_file=f.source_file,
                date=d, raw_product=f.product, raw_supplier=f.supplier,
                raw_restaurant=f.restaurant,
                quantity=f.quantity, unit_price=f.unit_price, total=f.total,
                product_master_id=pm_id, supplier_id=sup_id,
                restaurant_id=(rest.id if rest else None),
                accounting_system_id=(system.id if system else None),
            )
            db.add(pf)
            db.flush()
            total_facts += 1

            # 5) Сразу матчим с Топ-2
            dev = Deviation(purchase_fact_id=pf.id)

            if is_internal:
                dev.status = "internal"
            elif pm_id is None:
                dev.status = "unmapped_product"
                # лог в unmapped_items
                uid = db.execute(
                    select(UnmappedItem).filter_by(
                        raw_name=f.product, source=f.source,
                        system_id=(system.id if system else None),
                    )
                ).scalar_one_or_none()
                if uid:
                    uid.occurrence_count += 1
                    uid.last_seen = datetime.utcnow()
                else:
                    db.add(UnmappedItem(
                        raw_name=f.product, source=f.source,
                        system_id=(system.id if system else None),
                    ))
            else:
                # Топ-2 НА ДАТУ ФАКТА (а не текущий!)
                t = top2_on_date(pm_id, d)
                if t[0] is None:
                    dev.status = "no_quotes"
                elif t[2] is None:
                    dev.status = "no_top2"
                    dev.top1_supplier_id, dev.top1_price = t[0], t[1]
                else:
                    dev.top1_supplier_id, dev.top1_price = t[0], t[1]
                    dev.top2_supplier_id, dev.top2_price = t[2], t[3]
                    delta = f.unit_price - t[3]
                    delta_pct = (delta / t[3]) * 100 if t[3] else 0
                    overpay = max(0, delta) * f.quantity
                    dev.delta_per_unit = delta
                    dev.delta_pct = delta_pct
                    dev.overpayment = overpay
                    if delta <= 0:
                        dev.status = "green_top2"
                    elif delta_pct < 15:
                        dev.status = "yellow"
                    else:
                        dev.status = "red"

            # Восстанавливаем причину если она была указана раньше
            key = (d.isoformat(), f.product, f.supplier, f.restaurant or "")
            if key in saved_reasons:
                dev.reason_text, dev.reason_category = saved_reasons[key]

            db.add(dev)
            bucket[dev.status] = bucket.get(dev.status, 0) + 1

    return {"total_facts": total_facts, "buckets": bucket, "reasons_restored": sum(1 for _ in saved_reasons)}


# ============ MAIN ============

def main():
    init_db()
    db = SessionLocal()
    run = ImportRun(source="full_import")
    db.add(run)
    db.flush()
    print("→ seed static")
    seed_static(db)

    master_path = DRIVE_DIR / MASTER_FILE
    if master_path.exists():
        n = import_master_matrix(db, master_path)
        print(f"→ master matrix: {n} строк продуктов")

    mapping_path = DRIVE_DIR / MAPPING_FILE
    if mapping_path.exists():
        r = import_mapping(db, mapping_path)
        print(f"→ mapping: {r}")

    matrices = discover_supplier_matrices(DRIVE_DIR)
    print(f"→ найдено матриц поставщиков: {len(matrices)}")
    seen_supplier_norms: set[str] = set()
    for sup_name, path in matrices:
        seen_supplier_norms.add(normalize(sup_name))
        r = import_supplier_matrix(db, path, sup_name)
        print(f"→ matrix {sup_name}: {r}")

    # === ЧИСТКА ФАНТОМНЫХ ПОСТАВЩИКОВ ===
    # Поставщик в БД для которого больше нет xlsx-файла → фантом (файл удалён/переименован).
    # Сносим вместе с его quotes и changes. PriceHistory оставляем как исторический архив.
    from backend.models import PriceQuote as _PQ, PriceChange as _PC
    phantoms = [
        s for s in db.execute(select(Supplier)).scalars().all()
        if s.name_normalized not in seen_supplier_norms
    ]
    if phantoms:
        for ph in phantoms:
            db.execute(delete(_PQ).where(_PQ.supplier_id == ph.id))
            db.execute(delete(_PC).where(_PC.supplier_id == ph.id))
            db.delete(ph)
        print(f"→ удалено фантомных поставщиков: {len(phantoms)} ({[p.name for p in phantoms]})")

    print("→ facts:")
    r = import_facts(db)
    print(f"   {r}")

    run.status = "ok"
    run.finished_at = datetime.utcnow()
    db.commit()

    print("\n✅ Импорт завершён. БД: backend/data.db")


if __name__ == "__main__":
    main()
