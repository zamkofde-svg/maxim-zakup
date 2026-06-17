"""
Парсер выгрузки iiko: «Отчёт о закупочных ценах» (xlsx).

Формат:
  r1-r3: заголовки отчёта
  r4: шапка таблицы: Группа | Товар | Дата | Поставщик | Цена с НДС | Цена за ед. | Сумма | (пусто) | Количество
  r5+: данные. Группа и Товар идут «сверху вниз» — в первой строке группы заполнено,
       дальше None пока та же группа/товар. Строки «...Всего» — итоги по товару, пропускаем.

Возвращает список PurchaseFact.
"""
from __future__ import annotations
import re
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook


@dataclass
class PurchaseFact:
    source: str            # 'iiko' | 'storehouse'
    source_file: str       # имя файла
    group: str | None      # категория из выгрузки iiko (АРХИВ пф, БАКАЛЕЙНЫЕ ТОВАРЫ и т.д.)
    product: str           # название товара (как в выгрузке, может с *)
    date: str              # YYYY-MM-DD
    supplier: str          # как в выгрузке
    restaurant: str | None # получатель (только в StoreHouse-выгрузке)
    quantity: float
    unit_price: float      # цена за единицу с НДС
    total: float           # сумма с НДС
    row: int               # для дебага


def _norm(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _parse_iiko_purchase_report(ws, path: Path) -> list[PurchaseFact]:
    """Парсер для нового формата iiko «Отчёт о закупочных ценах».

    Шапка на ~4-й строке: Товар | Артикул | Дата и время | Поставщик |
    Номер документа | Цена с НДС, р. | Фасовка.

    Группа и количество отсутствуют — в этом отчёте только закупочные цены.
    quantity = 1.0 (плейсхолдер), чтобы матчинг по дате+поставщику+товару отработал.
    """
    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        a = _norm(ws.cell(r, 1).value)
        b = _norm(ws.cell(r, 2).value)
        if a == "Товар" and b in ("Артикул", "Дата и время"):
            header_row = r
            break
    if not header_row:
        return []  # формат не подошёл

    # Определим какие индексы у нужных колонок (может прыгать между версиями iiko)
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = _norm(ws.cell(header_row, c).value)
        if h:
            headers[h.lower()] = c
    col_product = headers.get("товар")
    col_date = headers.get("дата и время") or headers.get("дата")
    col_supplier = headers.get("поставщик")
    col_price = next((v for k, v in headers.items() if "цена" in k and "ндс" in k), None) \
                or headers.get("цена")
    col_qty = headers.get("количество") or headers.get("кол-во")
    col_pkg = headers.get("фасовка")

    facts: list[PurchaseFact] = []
    for r in range(header_row + 1, ws.max_row + 1):
        prod = _norm(ws.cell(r, col_product).value) if col_product else None
        if not prod:
            continue
        if prod.endswith("Всего") or prod.endswith("Итого"):
            continue
        d_val = ws.cell(r, col_date).value if col_date else None
        if not isinstance(d_val, datetime):
            # Может быть строка вроде "16.06.2026 13:30"
            try:
                d_val = datetime.strptime(str(d_val).split()[0], "%d.%m.%Y")
            except Exception:
                continue
        sup = _norm(ws.cell(r, col_supplier).value) if col_supplier else None
        if not sup:
            continue
        try:
            price = float(ws.cell(r, col_price).value) if col_price else None
        except (TypeError, ValueError):
            continue
        if price is None or price <= 0:
            continue
        qty_val = ws.cell(r, col_qty).value if col_qty else None
        try:
            qty = float(qty_val) if qty_val not in (None, "") else 1.0
        except (TypeError, ValueError):
            qty = 1.0
        facts.append(PurchaseFact(
            source="iiko",
            source_file=path.name,
            group=None,
            product=prod,
            date=d_val.strftime("%Y-%m-%d"),
            supplier=sup,
            restaurant=None,
            quantity=qty,
            unit_price=price,
            total=qty * price,
            row=r,
        ))
    return facts


def parse_iiko(path: str | Path) -> list[PurchaseFact]:
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active  # обычно один лист

    # Найти строку шапки (содержит «Группа» в A и «Товар» в B) — старый формат
    header_row = None
    for r in range(1, min(ws.max_row, 20) + 1):
        if _norm(ws.cell(r, 1).value) == "Группа" and _norm(ws.cell(r, 2).value) == "Товар":
            header_row = r
            break
    if not header_row:
        # Может быть новый формат «Отчёт о закупочных ценах»: Товар | Артикул | Дата | ...
        v2 = _parse_iiko_purchase_report(ws, path)
        if v2:
            return v2
        raise ValueError(f"Не нашёл шапку iiko ни в одном из форматов: {path}")

    facts: list[PurchaseFact] = []
    current_group: str | None = None
    current_product: str | None = None

    for r in range(header_row + 1, ws.max_row + 1):
        a = _norm(ws.cell(r, 1).value)
        b = _norm(ws.cell(r, 2).value)
        c = ws.cell(r, 3).value  # дата
        d = _norm(ws.cell(r, 4).value)  # поставщик
        e = ws.cell(r, 5).value  # цена с НДС
        f = ws.cell(r, 6).value  # цена за ед.
        g = ws.cell(r, 7).value  # сумма
        i = ws.cell(r, 9).value  # количество

        # Обновление текущей группы/товара
        if a:
            current_group = a
        if b:
            # Строка «...Всего» — итог, пропускаем
            if b.endswith("Всего") or b.endswith("Итого"):
                # сбросим current_product чтобы не приклеить следующий ряд к итогу
                continue
            current_product = b

        # Это строка-документ: должна быть дата + поставщик + количество
        if not isinstance(c, datetime) or not d or i is None:
            continue
        if current_product is None:
            continue

        try:
            qty = float(i)
            unit_price = float(f) if f is not None else float(e)
            total = float(g) if g is not None else qty * unit_price
        except (TypeError, ValueError):
            continue

        facts.append(PurchaseFact(
            source="iiko",
            source_file=path.name,
            group=current_group,
            product=current_product,
            date=c.strftime("%Y-%m-%d"),
            supplier=d,
            restaurant=None,
            quantity=qty,
            unit_price=unit_price,
            total=total,
            row=r,
        ))

    return facts


def main():
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "../sample-data/iiko_or_sh_2.xlsx"
    facts = parse_iiko(path)
    print(f"=== iiko: {path} ===")
    print(f"Записей закупок: {len(facts)}")

    # Топ-10 примеров
    print("\nПервые 5 строк:")
    for f in facts[:5]:
        print(f"  {f.date} | {f.product[:50]:<52} | {f.supplier[:35]:<37} | "
              f"кол.{f.quantity:>7} | цена {f.unit_price:>8.2f} | сумма {f.total:>10.2f}")

    # Статистика
    from collections import Counter
    suppliers = Counter(f.supplier for f in facts)
    groups = Counter(f.group for f in facts if f.group)
    products = Counter(f.product for f in facts)
    print(f"\nУникальных товаров: {len(products)}")
    print(f"Уникальных поставщиков: {len(suppliers)}")
    print(f"Уникальных групп: {len(groups)}")
    print(f"\nТоп-5 поставщиков по числу позиций:")
    for s, n in suppliers.most_common(5):
        print(f"  {n:>4} × {s}")
    print(f"\nТоп-5 групп по числу позиций:")
    for g, n in groups.most_common(5):
        print(f"  {n:>4} × {g}")


if __name__ == "__main__":
    main()
