"""
Распространение мастер-матрицы на матрицы всех поставщиков.

Действия (АДДИТИВНЫЕ — никогда не удаляем):
- Если у поставщика нет вкладки, которая есть в мастере → создаём её с шапкой
- Если в существующей вкладке нет позиции из мастера → дописываем строку

Цены поставщика никогда не трогаются. Лишние позиции у поставщика не удаляются.

Используется при:
- Демо: добавил «тест1» в мастер → нажал кнопку → разлетелось по 8 матрицам
- Регулярно: можно повесить на cron или ручной вызов

Требует scope: drive (для list files) + spreadsheets (для read/write листов).
"""
from __future__ import annotations
import io
import json
import os
import re
from pathlib import Path
from typing import Optional
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

SA_PATH = Path.home() / ".config" / "maxim-zakup" / "sa.json"
SCOPES = [
    "https://www.googleapis.com/auth/drive",        # для list files и export
    "https://www.googleapis.com/auth/spreadsheets",  # для read/write листов
]

# Имена файлов которые НЕ матрицы поставщиков (служебные)
MASTER_FILENAME = "Матрица(для изменения позиций)"
NON_SUPPLIER_NAMES = {
    MASTER_FILENAME,
    "Карта сопоставлений", "Топ 2", "Сопоставление", "Сводная",
}
SUPPLIER_PREFIXES = ("ООО ", "АО ", "ИП ", "ПАО ", "ЗАО ")


def _load_credentials():
    env_content = os.environ.get("GOOGLE_SA_JSON_CONTENT")
    if env_content:
        info = json.loads(env_content)
        return service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    if SA_PATH.exists():
        return service_account.Credentials.from_service_account_file(str(SA_PATH), scopes=SCOPES)
    raise RuntimeError("Нет credentials: ни GOOGLE_SA_JSON_CONTENT env, ни файла sa.json")


def _norm(s) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).strip()).lower().replace('"', "").replace("«", "").replace("»", "")


def _get_services():
    creds = _load_credentials()
    return (
        build("drive", "v3", credentials=creds, cache_discovery=False),
        build("sheets", "v4", credentials=creds, cache_discovery=False),
    )


def _list_files(drive) -> list[dict]:
    """Все файлы, видимые service account."""
    resp = drive.files().list(
        pageSize=200,
        fields="files(id, name, mimeType)",
        corpora="allDrives",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        q="mimeType != 'application/vnd.google-apps.folder'",
    ).execute()
    return resp.get("files", [])


def _is_supplier_file(name: str) -> bool:
    if name in NON_SUPPLIER_NAMES:
        return False
    return any(name.startswith(p) for p in SUPPLIER_PREFIXES)


def _read_all_sheets(sheets, spreadsheet_id: str, sheet_titles: list[str]) -> dict[str, list[str]]:
    """ОДНИМ запросом читает A:A для всех вкладок. Возвращает {sheet_title: normalized_names}."""
    if not sheet_titles:
        return {}
    ranges = [f"'{t}'!A1:A1000" for t in sheet_titles]
    resp = sheets.spreadsheets().values().batchGet(
        spreadsheetId=spreadsheet_id, ranges=ranges
    ).execute()
    result = {}
    for vr in resp.get("valueRanges", []):
        # Range приходит обратно в формате 'Title'!A1:A1000
        range_str = vr.get("range", "")
        title = range_str.split("!")[0].strip("'")
        values = vr.get("values", [])
        names = []
        for row in values[1:]:  # skip header
            if not row:
                continue
            v = row[0] if row else None
            if not v:
                continue
            if str(v).strip().startswith("Наименование"):
                continue
            names.append(_norm(v))
        result[title] = names
    return result


def _download_file(drive, file_id: str, mime: str) -> bytes:
    """Скачивает Google Sheet или xlsx → байты xlsx."""
    if mime == "application/vnd.google-apps.spreadsheet":
        req = drive.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


def _read_master_from_drive(drive, sheets, drive_files: list[dict]) -> tuple[str | None, dict[str, tuple[list[str], list, list[list]]]]:
    """Из мастер-матрицы (xlsx ИЛИ Google Sheet) возвращает:
    - master_file_id (для логирования)
    - {sheet_name: (normalized_names, raw_header, all_rows_in_order)}
    """
    master = next((f for f in drive_files if f["name"] == MASTER_FILENAME), None)
    if not master:
        return None, {}

    # Скачиваем мастер как xlsx (через Drive — работает и для Excel, и для нативного Sheet)
    blob = _download_file(drive, master["id"], master["mimeType"])
    wb = load_workbook(io.BytesIO(blob), data_only=True)

    out = {}
    for title in wb.sheetnames:
        ws = wb[title]
        # Шапка из первой строки
        max_col = ws.max_column or 1
        header = [ws.cell(1, c).value or "" for c in range(1, max_col + 1)]

        names_norm = []
        original_rows = []  # для сохранения порядка
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if not v:
                continue
            name = str(v).strip()
            if not name or name.startswith("Наименование"):
                continue
            names_norm.append(_norm(name))
            original_rows.append([name])  # сохраняем только колонку A
        out[title] = (names_norm, header, original_rows)
    return master["id"], out


def sync(dry_run: bool = False) -> dict:
    """
    Главная функция. Возвращает план изменений.
    Если dry_run=True — ничего не пишет в Google Sheets, только показывает план.
    """
    drive, sheets = _get_services()
    files = _list_files(drive)
    master_id, master_data = _read_master_from_drive(drive, sheets, files)
    if not master_id:
        raise RuntimeError("Мастер-матрица не найдена в Drive")

    plan = {
        "supplier_changes": [],
        "dry_run": dry_run,
        "total_sheets_added": 0,
        "total_rows_added": 0,
    }

    supplier_files = [f for f in files if _is_supplier_file(f["name"]) and f["mimeType"] == "application/vnd.google-apps.spreadsheet"]

    for sup_file in supplier_files:
        sup_id = sup_file["id"]
        sup_name = sup_file["name"]

        # 1 запрос — метаданные вкладок поставщика
        meta = sheets.spreadsheets().get(spreadsheetId=sup_id, fields="sheets.properties").execute()
        sup_sheet_titles = {s["properties"]["title"] for s in meta["sheets"]}

        # 1 batch-запрос — все позиции по всем вкладкам которые есть И у поставщика, И в мастере
        common_titles = [t for t in master_data.keys() if t in sup_sheet_titles]
        sup_data = _read_all_sheets(sheets, sup_id, common_titles)

        sheets_added: list[str] = []
        rows_added: dict[str, list[str]] = {}  # {sheet: [names]}

        for master_sheet_title, (master_names_norm, master_header, master_rows) in master_data.items():
            if master_sheet_title not in sup_sheet_titles:
                # Нужно добавить вкладку с шапкой и всеми позициями
                if not dry_run:
                    sheets.spreadsheets().batchUpdate(
                        spreadsheetId=sup_id,
                        body={"requests": [{"addSheet": {"properties": {"title": master_sheet_title}}}]},
                    ).execute()
                    values_to_write = [master_header] + [[row[0]] for row in master_rows]
                    sheets.spreadsheets().values().update(
                        spreadsheetId=sup_id,
                        range=f"'{master_sheet_title}'!A1",
                        valueInputOption="USER_ENTERED",
                        body={"values": values_to_write},
                    ).execute()
                sheets_added.append(master_sheet_title)
                rows_added[master_sheet_title] = [row[0] for row in master_rows]
            else:
                sup_names_norm = sup_data.get(master_sheet_title, [])
                sup_set = set(sup_names_norm)

                missing = []
                for i, m_name in enumerate(master_names_norm):
                    if m_name not in sup_set:
                        missing.append(master_rows[i][0])

                if missing:
                    if not dry_run:
                        sheets.spreadsheets().values().append(
                            spreadsheetId=sup_id,
                            range=f"'{master_sheet_title}'!A:A",
                            valueInputOption="USER_ENTERED",
                            insertDataOption="INSERT_ROWS",
                            body={"values": [[name] for name in missing]},
                        ).execute()
                    rows_added[master_sheet_title] = missing

        if sheets_added or rows_added:
            plan["supplier_changes"].append({
                "supplier": sup_name,
                "sheets_added": sheets_added,
                "rows_added": rows_added,
                "rows_added_count": sum(len(v) for v in rows_added.values()),
            })
            plan["total_sheets_added"] += len(sheets_added)
            plan["total_rows_added"] += sum(len(v) for v in rows_added.values())

    return plan


def main():
    import sys
    dry = "--dry-run" in sys.argv
    print(f"== sync master → suppliers (dry_run={dry}) ==")
    result = sync(dry_run=dry)
    print(f"\nИтого:")
    print(f"  Вкладок добавлено:  {result['total_sheets_added']}")
    print(f"  Строк добавлено:    {result['total_rows_added']}")
    print(f"  Поставщиков затронуто: {len(result['supplier_changes'])}")
    for ch in result["supplier_changes"]:
        print(f"\n  → {ch['supplier']}")
        if ch["sheets_added"]:
            print(f"     создано вкладок: {ch['sheets_added']}")
        for sheet, names in ch["rows_added"].items():
            print(f"     в '{sheet}' добавлено {len(names)} строк: {names[:3]}{'...' if len(names) > 3 else ''}")


if __name__ == "__main__":
    main()
