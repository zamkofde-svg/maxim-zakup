"""
Сравнение мастер-матрицы с матрицами поставщиков.
- какие позиции есть в мастере, но НЕТ у поставщика (нужно добавить)
- какие позиции есть у поставщика, но НЕТ в мастере (лишние, или поставщик добавил сам)
"""
from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

DRIVE = Path(__file__).parent.parent / "sample-data" / "drive-sync"
MASTER = DRIVE / "Матрица(для изменения позиций).xlsx"

NON_SUPPLIER = {
    "Матрица(для изменения позиций).xlsx",
    "Карта сопоставлений.xlsx",
    "Топ 2.xlsx", "Сопоставление.xlsx", "Сводная.xlsx",
}
SUPPLIER_PREFIXES = ("ООО ", "АО ", "ИП ", "ПАО ", "ЗАО ")


def normalize(s):
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).strip()).lower().replace('"', "").replace("«", "").replace("»", "")


def read_positions(path: Path) -> dict[str, set[str]]:
    """{category_name: {normalized_position_name}}. Возвращает названия категорий из вкладок."""
    wb = load_workbook(path, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        positions = set()
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row, 1).value
            if not v:
                continue
            name = str(v).strip()
            if not name or name.startswith("Наименование"):
                continue
            positions.add(normalize(name))
        out[sheet] = positions
    return out


def main():
    print(f"=== МАСТЕР: {MASTER.name} ===")
    master = read_positions(MASTER)
    total_master = sum(len(v) for v in master.values())
    print(f"Вкладок: {len(master)}, позиций суммарно: {total_master}")
    for cat in sorted(master):
        print(f"  {cat:<30} {len(master[cat])} позиций")

    print(f"\n=== ПОСТАВЩИКИ ===\n")

    diff_summary = []
    for path in sorted(DRIVE.glob("*.xlsx")):
        if path.name in NON_SUPPLIER:
            continue
        if not any(path.name.startswith(p) for p in SUPPLIER_PREFIXES):
            continue
        sup_name = path.stem
        sup = read_positions(path)

        print(f"### {sup_name} ###")
        # Сравниваем категории
        master_cats = set(master.keys())
        sup_cats = set(sup.keys())

        same_cats = master_cats & sup_cats
        miss_cats_in_sup = master_cats - sup_cats
        extra_cats_in_sup = sup_cats - master_cats

        if miss_cats_in_sup:
            print(f"  ⚠ нет вкладок у поставщика: {sorted(miss_cats_in_sup)}")
        if extra_cats_in_sup:
            print(f"  ℹ лишние вкладки: {sorted(extra_cats_in_sup)}")

        # Считаем разницу по позициям в общих вкладках
        only_master = 0
        only_sup = 0
        common = 0
        for cat in sorted(same_cats):
            m_pos = master[cat]
            s_pos = sup[cat]
            only_master += len(m_pos - s_pos)
            only_sup += len(s_pos - m_pos)
            common += len(m_pos & s_pos)

        print(f"  Общие вкладки: {len(same_cats)}/{len(master_cats)}")
        print(f"  Позиций совпадают: {common}")
        print(f"  В мастере но нет у поставщика: {only_master}")
        print(f"  У поставщика но нет в мастере: {only_sup}")
        diff_summary.append((sup_name, only_master, only_sup, common))
        print()

    # Сводная таблица
    print("=== СВОДНАЯ ===")
    print(f"{'Поставщик':<35} {'нет у пост.':>12} {'лишних':>10} {'совпало':>10}")
    for name, a, b, c in diff_summary:
        print(f"{name:<35} {a:>12} {b:>10} {c:>10}")


if __name__ == "__main__":
    main()
