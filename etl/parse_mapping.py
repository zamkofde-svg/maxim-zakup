"""
Парсер карты сопоставлений.
Формат: одна Google Sheet с вкладками по категориям + вкладка «поставщики».

Колонки разные по категориям, но логика общая:
  A: название у поставщика (= то, что в матрице поставщика, колонка A)
  далее: 1..N колонок с именами в разных системах учёта (SH/ST, SH сыр/Chees, Техникум, Сорренто, ...)

Парсер:
  - читает шапку каждой вкладки → определяет имена систем учёта
  - для каждой строки выдаёт записи (supplier_name, system, system_name)
  - служебная вкладка «поставщики» обрабатывается отдельно — маппинг имён поставщиков
"""
from __future__ import annotations
import json
import re
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from openpyxl import load_workbook

# Вкладки, которые не являются категориями товаров
META_SHEETS = {"поставщики"}

# Нормализация имён систем учёта: разные синонимы → единый канонический ключ
SYSTEM_ALIASES = {
    "SH": "SH",
    "ST": "SH",                # ST = StoreHouse, синоним
    "SH сыр": "Chees",
    "Chees": "Chees",
    "CH": "Chees",
    "Техникум": "TEHNIKUM",
    "Сорренто": "Sorrento",
}


def _norm_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return re.sub(r"\s+", " ", s)


def _canonical_system(name: str | None) -> str | None:
    if not name:
        return None
    s = name.strip()
    return SYSTEM_ALIASES.get(s, s)  # неизвестные системы — оставляем как есть


@dataclass
class MappingRecord:
    category: str
    supplier_name: str       # "название как у поставщиков"
    system: str              # каноническое имя системы: SH / Chees / TEHNIKUM / Sorrento
    system_name: str         # имя позиции в этой системе


@dataclass
class SupplierAlias:
    canonical: str           # "АО Группа \"ЮФС\""
    alias: str               # "АО ГРУППА ЮФС"
    source: str              # из какой колонки (B / C / ...)


def _find_header_row(ws) -> tuple[int, dict[int, str]]:
    """
    Шапка может быть в r1 или r2 (в некоторых вкладках в r1 — только имена систем, в r2 — заголовок A).
    Возвращает (номер_строки_с_системами, dict {col_index: system_name}).
    Эвристика: ищем в первых 2 строках колонки, где значение похоже на имя системы учёта.
    """
    known_system_names = {"SH", "ST", "SH сыр", "Chees", "CH", "Техникум", "Сорренто"}

    for r in (1, 2):
        cols = {}
        for c in range(2, ws.max_column + 1):
            v = _norm_str(ws.cell(r, c).value)
            if v in known_system_names:
                cols[c] = v
        if cols:
            return r, cols
    return 1, {}


def parse_category_sheet(ws) -> list[MappingRecord]:
    header_row, system_cols = _find_header_row(ws)
    if not system_cols:
        return []  # вкладка без систем — пропускаем

    records: list[MappingRecord] = []
    start_row = header_row + 1
    # Иногда в r2 идёт текст «Наименование товара» — это просто подпись, пропускаем
    for r in range(start_row, ws.max_row + 1):
        supplier_name = _norm_str(ws.cell(r, 1).value)
        if not supplier_name or supplier_name == "Наименование товара" or supplier_name == "Наименование":
            continue

        for col, raw_system in system_cols.items():
            system_name = _norm_str(ws.cell(r, col).value)
            if not system_name:
                continue
            records.append(MappingRecord(
                category=ws.title,
                supplier_name=supplier_name,
                system=_canonical_system(raw_system),
                system_name=system_name,
            ))
    return records


def parse_suppliers_sheet(ws) -> list[SupplierAlias]:
    """Вкладка «поставщики»: A=каноническое, B/C — синонимы в разных системах."""
    aliases: list[SupplierAlias] = []
    for r in range(1, ws.max_row + 1):
        canonical = _norm_str(ws.cell(r, 1).value)
        if not canonical:
            continue
        for col in (2, 3):
            alias = _norm_str(ws.cell(r, col).value)
            if alias:
                aliases.append(SupplierAlias(
                    canonical=canonical,
                    alias=alias,
                    source=chr(ord('A') + col - 1),
                ))
    return aliases


def parse(path: str | Path) -> tuple[list[MappingRecord], list[SupplierAlias]]:
    wb = load_workbook(path, data_only=True)
    all_records: list[MappingRecord] = []
    suppliers: list[SupplierAlias] = []
    for name in wb.sheetnames:
        ws = wb[name]
        if name.lower() in {s.lower() for s in META_SHEETS}:
            suppliers = parse_suppliers_sheet(ws)
        else:
            all_records.extend(parse_category_sheet(ws))
    return all_records, suppliers


def main():
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "../sample-data/mapping.xlsx"
    records, suppliers = parse(path)

    print(f"=== Карта сопоставлений: {path} ===")
    print(f"Всего записей: {len(records)}")
    print(f"Маппинг поставщиков: {len(suppliers)}\n")

    # Статистика по категориям
    by_cat = defaultdict(lambda: defaultdict(int))
    for r in records:
        by_cat[r.category][r.system] += 1

    print(f"{'Категория':<25} {'Систем':>8} {'Всего записей':>15}  Распределение по системам")
    for cat in sorted(by_cat):
        systems = by_cat[cat]
        total = sum(systems.values())
        distrib = ", ".join(f"{s}:{n}" for s, n in sorted(systems.items()))
        print(f"  {cat:<23} {len(systems):>8} {total:>15}  {distrib}")

    # Системы учёта: всё что встретилось
    all_systems = sorted({r.system for r in records})
    print(f"\nСистемы учёта (канонические): {all_systems}")

    # Уникальные мастер-позиции (= уникальные supplier_name)
    unique_products = {r.supplier_name for r in records}
    print(f"Уникальных мастер-позиций (по supplier_name): {len(unique_products)}")


if __name__ == "__main__":
    main()
