"""
Детектор и чистка шаблонного мусора в матрицах поставщиков.

Проблема: часть поставщиков копируют шаблон матрицы друг у друга вместе с
числовыми значениями в колонке B ("Упаковка брутто"). В результате на одной
и той же позиции у 20+ поставщиков стоит одно и то же число в B, не имеющее
отношения к реальному весу упаковки. Парсер раньше принимал это за цену,
теперь не принимает — но мусор лучше всё же убрать, чтобы матрицы выглядели
честно и не путали людей.

Алгоритм:
  1) Скан всех xlsx в /data/drive-sync. Для каждой позиции (вкладка + имя)
     собираем значение колонки B у каждого поставщика.
  2) Из имени позиции извлекаем фасовку ('0,250 гр.' → 0.25 кг). Если у >=N
     поставщиков значение в B одинаковое И сильно не сходится с фасовкой
     (отличается в TOLERANCE раз и более) — это копипаста шаблона.
  3) В Drive у этих поставщиков очищаем ячейку B (через Sheets API
     values.batchClear с нормализацией имён вкладок).

Запуск:
    python -m etl.clean_template_garbage           # реальная чистка
    python -m etl.clean_template_garbage --dry     # только показать что найдёт

После чистки запусти `python -m etl.sync_from_drive && python -m backend.importer`
чтобы БД и Топ-2 подхватили изменения.
"""
from __future__ import annotations
import json
import re
import sys
from collections import defaultdict, Counter
from pathlib import Path
from openpyxl import load_workbook

# Импорт из соседних модулей
_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))
from sync_master_to_suppliers import _get_services, _list_files, _execute_with_retry, _title_key  # noqa

DRIVE_DIR = Path("/data/drive-sync")
SKIP_FILES = ("Топ", "Матр", "Сопост", "Свод", "Карт")
THRESHOLD_SUPPLIERS = 5  # сколько поставщиков должны иметь одинаковое значение
TOLERANCE = 5.0          # во сколько раз B должен отличаться от фасовки в имени


def _parse_pkg_kg(name: str | None) -> float | None:
    """Извлечь фасовку из имени и привести к килограммам. None если не нашли."""
    if not name:
        return None
    s = str(name).lower().replace(",", ".")
    for pat, factor in [
        (r"(\d+(?:\.\d+)?)\s*кг", 1.0),
        (r"(\d+(?:\.\d+)?)\s*kg", 1.0),
        (r"(\d+(?:\.\d+)?)\s*л(?![а-я])", 1.0),
        (r"(\d+(?:\.\d+)?)\s*мл", 0.001),
        (r"(\d+(?:\.\d+)?)\s*гр(?:амм)?", 0.001),
        (r"(\d+(?:\.\d+)?)\s*г(?![а-я])", 0.001),
        (r"(\d+(?:\.\d+)?)\s*шт", None),
    ]:
        m = re.search(pat, s)
        if m:
            if factor is None:
                return None
            try:
                return float(m.group(1)) * factor
            except ValueError:
                continue
    return None


def _to_float(v) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except Exception:
        return None


def detect() -> list[dict]:
    """Сканирует /data/drive-sync и возвращает список позиций-кандидатов на чистку."""
    positions: dict[tuple[str, str], dict[str, tuple[float, str]]] = defaultdict(dict)
    suppliers = []
    for f in sorted(DRIVE_DIR.glob("*.xlsx")):
        if any(s in f.name for s in SKIP_FILES):
            continue
        suppliers.append(f.stem)
        try:
            wb = load_workbook(f, data_only=True)
        except Exception:
            continue
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                raw = ws.cell(r, 1).value
                if not raw:
                    continue
                name = str(raw).strip()
                if not name or name.lower().startswith("наименование"):
                    continue
                v = _to_float(ws.cell(r, 2).value)
                if v is None or v <= 0:
                    continue
                positions[(sheet_name, name.lower())][f.stem] = (v, name)
    garbage = []
    for (sheet, lower_name), sup_to_data in positions.items():
        if not sup_to_data:
            continue
        any_sup = next(iter(sup_to_data.values()))
        raw_name = any_sup[1]
        expected = _parse_pkg_kg(raw_name)
        if expected is None or expected <= 0:
            continue
        by_value: dict[float, list[str]] = defaultdict(list)
        for sup, (b, _) in sup_to_data.items():
            by_value[b].append(sup)
        for b_val, sups in by_value.items():
            if len(sups) < THRESHOLD_SUPPLIERS:
                continue
            ratio = b_val / expected
            if ratio > TOLERANCE or ratio < 1.0 / TOLERANCE:
                garbage.append({
                    "sheet": sheet,
                    "name": raw_name,
                    "lower": lower_name,
                    "B": b_val,
                    "expected_kg_from_name": expected,
                    "count": len(sups),
                    "suppliers": sorted(set(sups)),
                })
    garbage.sort(key=lambda x: -x["count"])
    return garbage


def clean(garbage: list[dict]) -> dict:
    """Очищает ячейки B по списку. Возвращает {ячеек: N, ошибок: M}."""
    drive, sheets = _get_services()
    files = _list_files(drive)
    file_by_name = {f["name"]: f for f in files}
    affected = sorted({sup for g in garbage for sup in g["suppliers"]})

    def _norm_lower(s):
        return str(s).strip().lower() if s else ""

    total_cells = 0
    errors = 0
    for sup_name in affected:
        f = file_by_name.get(sup_name)
        if not f:
            print(f"  ✗ файл не найден: {sup_name}")
            errors += 1
            continue
        sup_targets = [g for g in garbage if sup_name in g["suppliers"]]
        if not sup_targets:
            continue
        try:
            meta = _execute_with_retry(
                sheets.spreadsheets().get(spreadsheetId=f["id"], fields="sheets.properties"),
                label=f"meta({sup_name})",
            )
        except Exception as e:
            print(f"  ✗ {sup_name}: meta failed: {e}")
            errors += 1
            continue
        real_titles = [s["properties"]["title"] for s in meta["sheets"]]
        real_by_key = {_title_key(t): t for t in real_titles}
        sheet_map: dict[str, str] = {}
        for g in sup_targets:
            rt = real_by_key.get(_title_key(g["sheet"]))
            if rt:
                sheet_map[g["sheet"]] = rt
        if not sheet_map:
            continue
        used_real = set(sheet_map.values())
        ranges = [f"'{t}'!A1:A2000" for t in used_real]
        try:
            resp = _execute_with_retry(
                sheets.spreadsheets().values().batchGet(spreadsheetId=f["id"], ranges=ranges),
                label=f"read({sup_name})",
            )
        except Exception as e:
            print(f"  ✗ {sup_name}: read failed: {e}")
            errors += 1
            continue
        row_by_sheet: dict[str, dict[str, int]] = {}
        for vr in resp.get("valueRanges", []):
            title = vr.get("range", "").split("!")[0].strip("'")
            row_by_sheet[title] = {}
            for idx, row in enumerate(vr.get("values", [])):
                if not row:
                    continue
                v = row[0] if row else None
                if v is None:
                    continue
                row_by_sheet[title][_norm_lower(v)] = idx + 1
        clear_ranges = []
        for g in sup_targets:
            real_t = sheet_map.get(g["sheet"])
            if not real_t:
                continue
            rn = row_by_sheet.get(real_t, {}).get(g["lower"])
            if rn is None:
                continue
            clear_ranges.append(f"'{real_t}'!B{rn}")
        if not clear_ranges:
            continue
        try:
            _execute_with_retry(
                sheets.spreadsheets().values().batchClear(
                    spreadsheetId=f["id"], body={"ranges": clear_ranges},
                ),
                label=f"clear({sup_name})",
            )
            total_cells += len(clear_ranges)
            print(f"  ✓ {sup_name}: -{len(clear_ranges)}")
        except Exception as e:
            print(f"  ✗ {sup_name}: clear failed: {e}")
            errors += 1
    return {"cells_cleared": total_cells, "errors": errors}


def main():
    dry = "--dry" in sys.argv or "--dry-run" in sys.argv
    garbage = detect()
    print(f"Подозрительных позиций: {len(garbage)}")
    by_sheet = Counter(g["sheet"] for g in garbage)
    for sheet, n in by_sheet.most_common():
        print(f"  [{sheet}]: {n}")
    print()
    for g in garbage[:30]:
        print(f"  [{g['sheet']:18s}] '{g['name'][:55]:55s}' B={g['B']:>8}  vs  {g['expected_kg_from_name']:.3f} кг "
              f"(у {g['count']} пост.)")
    if len(garbage) > 30:
        print(f"  ... и ещё {len(garbage) - 30}")
    if dry:
        print("\n--dry: чистка не запускалась.")
        return
    print(f"\nЧистка...")
    res = clean(garbage)
    print(f"\nИТОГО: ячеек очищено {res['cells_cleared']}, ошибок {res['errors']}")


if __name__ == "__main__":
    main()
