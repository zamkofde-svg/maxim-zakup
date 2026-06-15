"""
Парсер матрицы поставщика (Google Sheets / xlsx).

Вход: путь к xlsx-файлу (один поставщик)
Выход: список dict с полями: supplier, category, product, unit_type, unit_price, pkg_net, pkg_gross

Формат матрицы (по реальному шаблону «Максим»):
  - Вкладки = категории (Сыры, Молочка, ...)
  - Колонки:
    A=Наименование товара, B=Упаковка брутто, C=Стоимость (формула),
    D=Упаковка нетто, E=Цена за кг/литр (Сыры/Молочка) ИЛИ Цена упаковка
"""
from __future__ import annotations
import json
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from openpyxl import load_workbook

# Старый хардкод оставлен как fallback на случай если в шапке нет слова "Цена".
# В норме unit_type определяется по тексту самого заголовка колонки.
UNIT_PRICE_CATEGORIES = {"Сыры", "Молочка"}


def _norm_str(v) -> str | None:
    """Нормализация строки: trim, схлопывание пробелов."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # схлопываем подряд идущие пробелы
    s = re.sub(r"\s+", " ", s)
    return s


def _to_float(v) -> float | None:
    """Безопасное приведение к float, понимает '12,34' и '12.34', пропускает мусор."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


@dataclass
class PriceQuote:
    supplier: str
    category: str
    product: str        # очищенное название
    product_raw: str    # исходное название (для дебага)
    unit_type: str      # 'kg_or_l' | 'pkg'
    unit_price: float
    pkg_net: float | None
    pkg_gross: float | None
    row: int            # номер строки в файле — для дебага
    supplier_comment: str | None = None  # свободный текст из колонки «Комментарий» матрицы поставщика


def _header_keyword_score(title: str) -> int:
    """Чем больше score, тем уверенней эта колонка содержит цены.
    0 = «Цена» в заголовке нет вовсе.
    """
    if not title:
        return 0
    t = str(title).lower()
    if "цена" not in t and "стоимост" not in t and "price" not in t and "₽" not in t and "руб" not in t:
        return 0
    # «Цена кг/литр» / «Цена за кг» — самая правильная финальная цена за единицу
    if "кг" in t or "литр" in t or "/л" in t or "за кг" in t:
        return 30
    # «Цена упаковка» / «Цена (упаковка)»
    if "упак" in t:
        return 20
    # просто «Цена» / «Стоимость» / «Price» / «₽»
    return 10


def _detect_price_column_by_header(ws) -> tuple[int | None, str]:
    """Возвращает (номер колонки 1-based, текст заголовка) на основании ШАПКИ.
    Если в шапке нет ни одной колонки со словом «Цена/Стоимость» — возвращает (None, '').
    """
    best_col, best_score, best_title = None, 0, ""
    for c in range(2, 8):  # B..G
        title = ws.cell(1, c).value
        score = _header_keyword_score(title)
        if score > best_score:
            best_col, best_score, best_title = c, score, str(title or "")
    return best_col, best_title


def _list_price_columns_by_header(ws) -> list[tuple[int, str, int]]:
    """Возвращает все колонки шапки которые похожи на «Цена/Стоимость»,
    в виде [(col_index_1based, header_text, score), ...] отсортированно по score↓.
    Первая = самая «правильная» (Цена кг/литр), последние = просто «Цена».
    Если ни одной — пустой список.
    """
    out = []
    for c in range(2, 8):
        title = ws.cell(1, c).value
        score = _header_keyword_score(title)
        if score > 0:
            out.append((c, str(title or ""), score))
    out.sort(key=lambda x: -x[2])
    return out


def _unit_type_for_header(header_text: str, category: str) -> str:
    """Определяем unit_type по тексту шапки. Если в шапке не было намёков —
    fallback на список категорий (исторический хардкод)."""
    t = (header_text or "").lower()
    if "кг" in t or "литр" in t or "/л" in t:
        return "kg_or_l"
    if "упак" in t:
        return "pkg"
    return "kg_or_l" if category in UNIT_PRICE_CATEGORIES else "pkg"


def _detect_price_column_heuristic(ws) -> int | None:
    """Fallback: если в шапке нет «Цена» (упрощённые матрицы вроде Овощифруктов
    у Пашаяна), берём колонку B..F с максимальным числом положительных значений.
    Это даёт хоть какие-то цены для тех редких вкладок, у которых нет шапки.
    """
    counts = {c: 0 for c in range(2, 7)}
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not a:
            continue
        s = str(a).strip()
        if not s or s.startswith("Наименование"):
            continue
        for c in range(2, 7):
            v = ws.cell(r, c).value
            if v in (None, ""):
                continue
            try:
                fv = float(str(v).replace(" ", "").replace(",", "."))
            except (TypeError, ValueError):
                continue
            if fv > 0:
                counts[c] += 1
    if max(counts.values(), default=0) == 0:
        return None
    # Эвристика менее строгая: даём шанс C (типовой для упрощённых матриц), потом E
    priority = [3, 5, 2, 4, 6]
    best = max(counts.values())
    for c in priority:
        if counts[c] == best:
            return c
    return None


def _detect_comment_column(ws) -> int | None:
    """Ищет колонку с заголовком «Комментарий»/«Примечание»/«Comment». 1-based."""
    for c in range(2, 10):  # B..I
        title = ws.cell(1, c).value
        if title is None: continue
        t = str(title).lower()
        if "коммент" in t or "примеч" in t or "comment" in t:
            return c
    return None


def parse(path: str | Path, supplier_name: str) -> list[PriceQuote]:
    wb = load_workbook(path, data_only=True)  # читаем вычисленные значения формул
    quotes: list[PriceQuote] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        category = sheet_name.strip()

        # Все колонки шапки которые содержат «Цена/Стоимость», отсортированные по «правильности»:
        # сначала «Цена кг/литр», потом «Цена упаковка», потом просто «Цена».
        # Для каждой строки берём ПЕРВУЮ непустую — так подхватим и тех поставщиков
        # которые заполнили цену за упаковку, и тех кто заполнил цену за кг.
        price_cols = _list_price_columns_by_header(ws)
        comment_col = _detect_comment_column(ws)

        if not price_cols:
            # Шапки нет вовсе (например упрощённый Овощифрукты у Пашаяна: A=Наимен., B=Ед.изм).
            # Угадываем колонку по эвристике на всю вкладку.
            heur = _detect_price_column_heuristic(ws)
            if heur is None:
                continue
            price_cols = [(heur, "", 0)]

        for row in range(2, ws.max_row + 1):
            product_raw = ws.cell(row, 1).value
            product = _norm_str(product_raw)
            if not product:
                continue
            if product.startswith("Наименование"):
                continue

            # Перебираем колонки в порядке приоритета, берём первую с валидной ценой
            chosen_col = None
            chosen_header = ""
            chosen_price = None
            for col_idx, header_text, _ in price_cols:
                v = _to_float(ws.cell(row, col_idx).value)
                if v is None or v <= 0:
                    continue
                chosen_col = col_idx
                chosen_header = header_text
                chosen_price = v
                break

            if chosen_price is None:
                # Ни в одной из «ценовых» колонок числа нет — позиции у этого
                # поставщика нет. Не угадываем по другим колонкам.
                continue

            unit_type = _unit_type_for_header(chosen_header, category)

            # pkg_net/pkg_gross — берём только если их колонки не пересеклись с ценой
            pkg_gross = _to_float(ws.cell(row, 2).value) if chosen_col != 2 else None
            pkg_net = _to_float(ws.cell(row, 4).value) if chosen_col != 4 else None

            supplier_comment = None
            if comment_col is not None:
                v = ws.cell(row, comment_col).value
                if v is not None:
                    s = str(v).strip()
                    if s:
                        supplier_comment = s

            quotes.append(PriceQuote(
                supplier=supplier_name,
                category=category,
                product=product,
                product_raw=str(product_raw),
                unit_type=unit_type,
                unit_price=chosen_price,
                pkg_net=pkg_net,
                pkg_gross=pkg_gross,
                row=row,
                supplier_comment=supplier_comment,
            ))

    return quotes


def main():
    import sys
    if len(sys.argv) < 3:
        print("Usage: parse_supplier_matrix.py <xlsx> <supplier_name>")
        sys.exit(1)
    quotes = parse(sys.argv[1], sys.argv[2])
    out = [asdict(q) for q in quotes]
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
